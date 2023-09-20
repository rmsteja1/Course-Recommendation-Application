from flask_mail  import Message
from application import app
from flask import request, make_response, jsonify
from datetime import datetime
from application import db
from application import mail
import random
import math 
from passlib.hash import sha256_crypt
import bson.json_util as json_util
import openpyxl
import jwt
from functools import wraps
from datetime import datetime, timedelta
from uuid import uuid4
from flask_cors import cross_origin
from courseDict import courseDictNew as courseDictionary
from courseCodeDict import courseCodesDict as courseCodesDictionary
import joblib
import pandas as pd
from functions import courseAvailabilityPredictor
from functions import seCoursePredictionConditional
from functions import seCoursePredictionNonConditional
import boto3
import os

@app.route("/")
def index():
    return "Hello"

    
def downloadFile():
    try:
        session = boto3.Session(aws_access_key_id='AKIAYA2ZJPDIKNGR3VB6',aws_secret_access_key='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',)

    # Create an S3 client
        s3_client = session.client('s3')

    # Define the S3 bucket name and file name
        bucket_name = '295bmlmodel'
        file_name = 'rf_reg_new1.pkl'

    # Specify the local file path where you want to save the downloaded file
        current_directory = os.getcwd()

# Construct the local file path using the current directory
        filePath = "\\application\\"+file_name
        print(current_directory)
        current_directory = os.path.join(current_directory,"application")
        print(current_directory)
        local_file_path = os.path.join(current_directory,file_name)
        print(local_file_path)
    # Download the file
        s3_client.download_file(bucket_name, file_name, local_file_path)

        return "File Downloaded Successfully"
    except Exception as e:
        print(e)
        return "download failed"
    

@app.route("/downloadPickle")
def downloadPickleFile():
    return downloadFile()


def token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = None
        # jwt is passed in the request header
        if 'x-access-token' in request.headers:
            token = request.headers['x-access-token']
        # return 401 if token is not passed
        if not token:
            return jsonify({'message' : 'Token is missing !!'}), 401
  
        try:
            # decoding the payload to fetch the stored details
            data = jwt.decode(token, app.config['SECRET_KEY'],algorithms=['HS256'])
            current_user = db.users.find_one({"email": data["email"]})
            if(current_user == None):
                return jsonify({
                'message' : 'Token is invalid !!'
            }), 401
        except Exception as e:
            print(e)
            return jsonify({
                'message' : 'Token is invalid !!'
            }), 401
        # returns the current logged in users context to the routes
        return  f(current_user, *args, **kwargs)
  
    return decorated


def sendMail(OTP,Email):
    try:
        msg = Message('Verification of your account', sender =  'courseplanner@sjsu.edu', recipients = [Email])
        msg.body = "Hey User, \n \n OTP for verification is "+ OTP
        mail.send(msg)
    except Exception as e:
        print(e)
        return"Message sent failed", 500
    return "Message sent!", 200

def generateOTP():
    digits = [i for i in range(0, 10)]
    random_str = ""
    for i in range(6):
        index = math.floor(random.random() * 10)

        random_str += str(digits[index])
    return random_str

@app.route('/api/v1/register',methods = ["POST"])
@cross_origin(support_credentials=True)
def addUser():
    content_type = request.headers.get('Content-Type')
    if (content_type == 'application/json'):
        json = request.json
        try:
            user = db.users.find_one({"email":json["email"]})
            if user is not None:
                return "Email is already in use", 500
            now = datetime.utcnow()
            json["timestamp"]=now
            json["OTP"]=generateOTP()
            json["verified"]=False
            password = sha256_crypt.encrypt(json["password"])
            json["password"]=password
            uuid= uuid4()
            json["ID"] = str(uuid)
            sendMail(json["OTP"],json["email"])
            db.users.insert_one(json)
            return "Insertion Successful", 200
        except Exception as e :
            print(e)
            return 'Insertion Failed', 500
        
@app.route('/api/v1/verifyAccount', methods = ["POST"])
@cross_origin(support_credentials=True)
def verifyUser():
    json = request.json
    email = json["email"]
    OTP = json["OTP"]
    user = db.users.find_one({"email":email})
    if user == None:
        return "User not found", 404
    now = datetime.utcnow()
    then = user["timestamp"]
    duration = now - then
    duration_in_s = duration.total_seconds()
    minutes = divmod(duration_in_s, 60)[0] 
    if user["OTP"]==OTP and minutes<=15:
        myquery = { "email": email }
        newvalues = { "$set": { "verified": True } }
        db.users.update_one(myquery,newvalues)
        return "Account Verified", 200
    else:
        if(minutes<=15):
            return "Invalid OTP", 500
        else:
            return "OTP expired. Please generate new OTP", 500
        
@app.route('/api/v1/user', methods=["GET"])
@cross_origin(support_credentials=True)
@token_required
def getUser(currentUser):
    del currentUser["timestamp"]
    del currentUser["OTP"]
    del currentUser["password"]
    del currentUser["verified"]
    del currentUser["_id"]
    return json_util.dumps(currentUser),200

@app.route('/api/v1/login', methods=["POST"])
@cross_origin(support_credentials=True)
def login():
    json = request.json
    email = json["email"]
    password = json["password"]
    user = db.users.find_one({"email":email})
    if user == None:
        return "Invalid Credentials", 404
    if user["verified"] == False:
        return "Please verify your account.", 500
    if sha256_crypt.verify(password,user["password"]):
        token = jwt.encode({
            'email': user["email"],
            'exp' : datetime.utcnow() + timedelta(minutes = 30)
        }, app.config['SECRET_KEY'], algorithm='HS256')    
           
        return make_response(jsonify({'token' : token}), 201)
    else:
        return "Invalid Credentials", 500
    
@app.route('/api/v1/resendOTP', methods=["POST"])
@cross_origin(support_credentials=True)
def resendOTP():
    json = request.json
    email = json["email"]
    user = db.users.find_one({"email":email})
    if user == None:
        return "User not registered", 404
    if user["verified"] == True:
        return "User already verified their account", 401
    otp = generateOTP()
    now = datetime.utcnow()
    myquery = { "email": email }
    newvalues = { "$set": { "OTP": otp , "timestamp":now} }
    db.users.update_one(myquery,newvalues)
    sendMail(otp,email)
    return "OTP is sent successfully.", 200

@app.route('/api/v1/datapush', methods=["POST"])
@cross_origin(support_credentials=True)
def push():
    json = request.json
    db.degreerequirement.insert_one(json)
    return "Insertion Sucessful"

@app.route('/api/v1/pushrecdata', methods=["POST"])
def pushHistory():
    json = request.json
    db.recommendationhistory.insert_one(json)
    return "Insertion Sucessful"

@app.route('/retrieveData', methods=["GET"])
def retrieve():
    wrkbk = openpyxl.load_workbook("CS_100.xlsx")
  
    sh = wrkbk.active
    attr = ["coursecode","coursename","department","elective"]
# iterate through excel and display data
    for row in sh.iter_rows(min_row=2, min_col=1, max_row=sh.max_row, max_col=4):
        data = {}
        for i in range(len(row)):
            val = row[i].value
            if(isinstance(val,str)):
                data[attr[i]]=val.strip()
            data[attr[i]]=val
        db.coursecatalog.insert_one(data)
    #db.coursecatalog.delete_many({"elective":1})
    return "Retrieved"


@app.route("/api/v1/semesterPredict",methods=["GET"])
@cross_origin(support_credentials=True)
def predict():
    try:
        courseCodeString=request.args.get('courseCodes')
        courseCodeArray=courseCodeString.split(',')
        semester=request.args.get('semester')
        year=request.args.get('year')
        output=[]
        for course in courseCodeArray:
            skiplList=["CMPE 295A","CMPE 295B","CMPE 299A","CMPE 299B","CMPE 180A","CMPE 180B","CMPE 180C"]
            if course in skiplList:
                    output.append(1.0)
            else:
                prediction=courseAvailabilityPredictor(course,year,semester)
                output.append(prediction)
        return jsonify(output) , 200
    except Exception as e :
        print(e)

@app.route("/api/v1/degreePredict",methods=["GET"])
@cross_origin(support_credentials=True)
def completeCoursePrediction():
    department=request.args.get('department')
    specialization=request.args.get('specialization')
    startingSemester=request.args.get('startingSemester')
    startingYear=request.args.get('startingYear')
    semesterCount=request.args.get('semesterCount')
    crossSpecializations=request.args.get('crossSpecializations')
    electives=request.args.get('electives')
    conditionalAdmit=request.args.get('conditionalAdmit')

    if department=='Software Engineering' and conditionalAdmit=="yes":
        result=seCoursePredictionConditional(specialization,startingSemester,startingYear,semesterCount,crossSpecializations,electives,conditionalAdmit)
        return jsonify(result),200
    elif department=='Software Engineering' and conditionalAdmit=="no":
        result=seCoursePredictionNonConditional(specialization,startingSemester,startingYear,semesterCount,crossSpecializations,electives,conditionalAdmit)
        return jsonify(result),200
        
